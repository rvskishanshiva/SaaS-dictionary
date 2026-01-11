from openpyxl import load_workbook

f = "data/users.xlsx"

def user_exists(email):
    wb = load_workbook(f)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == email:
            return True
    return False


def add_user(name, email):
    if user_exists(email):
        return False
    wb = load_workbook(f)
    ws = wb.active
    ws.append([name, email])
    wb.save(f)
    return True


def display_users():
    wb = load_workbook(f)
    ws = wb.active
    print("\nName\tEmail")
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(row[0], "\t", row[1])
