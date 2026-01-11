from openpyxl import load_workbook

f = "data/history.xlsx"

def log_history(email, word):
    wb = load_workbook(f)
    ws = wb.active
    ws.append([email, word])
    wb.save(f)


def show_history(email):
    wb = load_workbook(f)
    ws = wb.active
    print("\n--- Search History ---")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == email:
            print(row[1])
