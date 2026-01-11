from openpyxl import load_workbook
from collections import Counter

f = "data/history.xlsx"

def top_searched_words():
    wb = load_workbook(f)
    ws = wb.active

    words = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        words.append(row[1])

    count = Counter(words)
    print("\n--- Top Searched Words ---")
    for word, n in count.most_common():
        print(word, "→", n)
