from openpyxl import load_workbook

f = "data/words.xlsx"

def search_word(word):
    wb = load_workbook(f)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == word:
            return row[1]
    return None


def add_word(word, meaning):
    wb = load_workbook(f)
    ws = wb.active
    ws.append([word, meaning])
    wb.save(f)
